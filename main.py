import shoppingSite_getInfo as scraping
import sys
import openpyxl
args = sys.argv
print(len(args))
if len(args) == 7:
    filename = args[1]
    sheetname = args[2]
    name_column = int(args[3])
    price_column = int(args[4])
    url_column = int(args[5])
    order_num_column = int(args[6])
    print(filename,sheetname,name_column,price_column,price_column,url_column,order_num_column)

    #print(scraping.get_name(url))
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    isSheet_END=True
    sheet_startrow =2
    sheet_url_array =list()
    sheet_order_num_array = list()
    sheet_row = sheet_startrow
    while isSheet_END:
        url = sheet.cell(row=sheet_row, column=url_column).value
        if url is not None:
            sheet_url_array.insert(sheet_row,url)
            sheet_row = sheet_row +1
        else:
            isSheet_END=False
    sheet_row = sheet_startrow
    isSheet_END=True
    while isSheet_END:
        order_num = sheet.cell(row=sheet_row, column=order_num_column).value
        if order_num is not None:
            sheet_order_num_array.insert(sheet_row,int(order_num))
            sheet_row = sheet_row +1
        else:
            isSheet_END=False

    sheet_row = sheet_startrow
    if len(sheet_order_num_array)==len(sheet_url_array):
        for i in range(len(sheet_url_array)):
            url = sheet_url_array[i]
            order_num= sheet_order_num_array[i]
            name = scraping.get_name(url)
            price = scraping.get_price(url,order_num)
            if(name is not None or price is not None):
                sheet.cell(row = sheet_row,column = name_column,value = name)
                sheet.cell(row = sheet_row,column = price_column,value = price)
            sheet_row= sheet_row+1
    else:
        print("error")
    workbook.save("sample3.xlsx")


