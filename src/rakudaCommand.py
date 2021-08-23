import shoppingSite_getInfo as scraping
import sys
import openpyxl
args = sys.argv

errorMessage_array=list()
def getErrorMessage():
    return errorMessage_array+ scraping.getErrorMessage()
def generateFile(filename,sheetname,name_column,order_num_column,price_column,url_column,saveFilename):
    global errorMessage_array
    errorMessage_array = list()
    scraping.clearErrorMessage()
    print(filename,sheetname,name_column,order_num_column,price_column,url_column,saveFilename)

    #print(scraping.get_name(url))
    try:
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook[sheetname]
        isSheet_END=True
        sheet_startrow =2
        sheet_url_array =list()
        sheet_order_num_array = list()
        sheet_row = sheet_startrow
        while isSheet_END:
            url = sheet.cell(row=sheet_row, column=url_column).value
            if url is not None:
                sheet_url_array.insert(sheet_row,url)
                sheet_row = sheet_row +1
            else:
                isSheet_END=False
        sheet_row = sheet_startrow
        isSheet_END=True
        while isSheet_END:
            order_num = sheet.cell(row=sheet_row, column=order_num_column).value
            if order_num is not None:
                sheet_order_num_array.insert(sheet_row,int(order_num))
                sheet_row = sheet_row +1
            else:
                isSheet_END=False

        sheet_row = sheet_startrow
        if len(sheet_order_num_array)==len(sheet_url_array):
            for i in range(len(sheet_url_array)):
                url = sheet_url_array[i]
                order_num= sheet_order_num_array[i]
                name = scraping.get_name(url)
                price = scraping.get_price(url,order_num)
                if(name is not None or price is not None):
                    sheet.cell(row = sheet_row,column = name_column,value = name)
                    sheet.cell(row = sheet_row,column = price_column,value = price)
                sheet_row= sheet_row+1
            workbook.save(saveFilename)
        else:
            errorMessage_array.append("Excelの注文数とURLの列の数が一致していません。")
    except KeyError:
        errorMessage_array.append(sheetname +"というシートは存在しません。")


    

if __name__=="__main__":
    if len(args) == 8:
        filename = args[1]
        sheetname = args[2]
        name_column = int(args[3])
        order_num_column = int(args[4])
        price_column = int(args[5])
        url_column = int(args[6])
        saveFilename = args[7]
        generateFile(filename,sheetname,name_column,order_num_column,price_column,url_column,saveFilename)